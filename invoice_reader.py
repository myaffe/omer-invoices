"""
Omer Invoices - reads invoices from Google Drive and extracts details to Excel.

Usage:
  python invoice_reader.py                    # process all files in the Drive folder
  python invoice_reader.py --subfolder march  # process files in a subfolder
  python invoice_reader.py --output report.xlsx
"""

import os
import io
import json
import base64
import argparse
import subprocess
import tempfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from pathlib import Path

import anthropic
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from dotenv import load_dotenv

load_dotenv()

SCOPES = ['https://www.googleapis.com/auth/drive']
SUPPORTED_TYPES = {
    'application/pdf',
    'image/jpeg',
    'image/png',
    'image/webp',
    'image/gif',
}

EXTRACTION_PROMPT = """You are an Israeli invoice/receipt data extractor. Analyze this invoice image and extract the following fields. The document is likely in Hebrew.

Return ONLY a valid JSON object with these fields:
{
  "vendor_name": "name of the business/vendor (in Hebrew if that's how it appears)",
  "vendor_id": "business number (ח.פ. / עוסק מורשה / ע.מ.) or empty string",
  "invoice_number": "invoice/receipt number or empty string",
  "invoice_date": "date in YYYY-MM-DD format",
  "description": "brief description of what was purchased/service provided (in Hebrew)",
  "amount_before_vat": number or null if not shown separately,
  "vat_amount": number or null if not shown separately,
  "total_amount": number (the final total paid),
  "currency": "ILS" or other currency code,
  "payment_method": "cash/credit/bank transfer/check or empty string if unknown",
  "notes": "any other relevant details (in Hebrew if applicable)"
}

CRITICAL accuracy rules:
- Read the vendor/business name EXACTLY as printed on the document header. Do NOT guess or approximate - copy the Hebrew text character by character
- The vendor name is usually at the TOP of the invoice in large/bold text, or next to the business number
- The business number (ח.פ./עוסק מורשה) is a 9-digit number - read it carefully
- All amounts should be numbers (not strings)
- If VAT is not shown separately, set amount_before_vat and vat_amount to null
- For the total, use the final amount including VAT (סה"כ לתשלום / סה"כ כולל מע"מ)
- For handwritten documents, read numbers very carefully - distinguish between 3 and 8, 0 and 6
- If you can't determine a field, use empty string or null
- Return ONLY the JSON, no other text"""


def get_drive_service():
    key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_KEY', '')
    if key_path and os.path.exists(key_path):
        creds = service_account.Credentials.from_service_account_file(key_path, scopes=SCOPES)
    else:
        key_json = os.environ.get('GOOGLE_SERVICE_ACCOUNT_KEY_JSON', '')
        if key_json:
            info = json.loads(key_json)
            creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
        else:
            raise ValueError("No Google service account credentials found")
    return build('drive', 'v3', credentials=creds)


def list_files(service, folder_id, subfolder=None):
    """List all supported files in the Drive folder."""
    if subfolder:
        # Find subfolder by name
        results = service.files().list(
            q=f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.folder' "
              f"and name contains '{subfolder}'",
            fields="files(id, name)"
        ).execute()
        folders = results.get('files', [])
        if not folders:
            print(f"Subfolder '{subfolder}' not found")
            return []
        folder_id = folders[0]['id']
        print(f"Using subfolder: {folders[0]['name']}")

    all_files = []
    page_token = None
    while True:
        results = service.files().list(
            q=f"'{folder_id}' in parents and mimeType != 'application/vnd.google-apps.folder'",
            fields="nextPageToken, files(id, name, mimeType, size, createdTime)",
            pageSize=100,
            pageToken=page_token
        ).execute()
        all_files.extend(results.get('files', []))
        page_token = results.get('nextPageToken')
        if not page_token:
            break

    supported = [f for f in all_files if f['mimeType'] in SUPPORTED_TYPES]
    print(f"Found {len(supported)} supported files ({len(all_files)} total)")
    return supported


def download_file(service, file_info):
    """Download a file from Drive and return bytes."""
    request = service.files().get_media(fileId=file_info['id'])
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


HEBREW_MONTHS = {
    1: 'ינואר', 2: 'פברואר', 3: 'מרץ', 4: 'אפריל',
    5: 'מאי', 6: 'יוני', 7: 'יולי', 8: 'אוגוסט',
    9: 'ספטמבר', 10: 'אוקטובר', 11: 'נובמבר', 12: 'דצמבר',
}


def get_or_create_month_folder(service, parent_folder_id):
    """Get or create a subfolder named after the current month (e.g. 'מרץ 2026')."""
    now = datetime.now()
    folder_name = f"{HEBREW_MONTHS[now.month]} {now.year}"

    # Check if folder already exists
    results = service.files().list(
        q=f"'{parent_folder_id}' in parents and mimeType='application/vnd.google-apps.folder' "
          f"and name='{folder_name}' and trashed=false",
        fields="files(id, name)"
    ).execute()
    folders = results.get('files', [])
    if folders:
        return folders[0]['id'], folder_name

    # Create the folder
    metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_folder_id],
    }
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder['id'], folder_name


def move_file_to_folder(service, file_id, target_folder_id):
    """Move a file to a target folder in Drive."""
    # Get current parents
    f = service.files().get(fileId=file_id, fields='parents').execute()
    previous_parents = ','.join(f.get('parents', []))
    # Move file
    service.files().update(
        fileId=file_id,
        addParents=target_folder_id,
        removeParents=previous_parents,
        fields='id, parents'
    ).execute()


def pdf_to_images(pdf_bytes):
    """Convert PDF pages to PNG images using pdftoppm."""
    images = []
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        # Start with 200 DPI, reduce if image is too large (5MB limit)
        for dpi in [200, 150, 100]:
            images = []
            with tempfile.TemporaryDirectory() as tmpdir:
                subprocess.run(
                    ['pdftoppm', '-png', '-r', str(dpi), tmp_path, f'{tmpdir}/page'],
                    capture_output=True, check=True
                )
                for page_file in sorted(Path(tmpdir).glob('page-*.png')):
                    img_data = page_file.read_bytes()
                    images.append(img_data)

            # Check if largest image is under 5MB
            if images and max(len(img) for img in images) < 4_500_000:
                break
    finally:
        os.unlink(tmp_path)

    return images


def extract_invoice_data(client, image_bytes, media_type, filename):
    """Send image to Claude and extract invoice data."""
    b64 = base64.standard_b64encode(image_bytes).decode('utf-8')

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": b64,
                    }
                },
                {
                    "type": "text",
                    "text": EXTRACTION_PROMPT
                }
            ]
        }]
    )

    text = response.content[0].text.strip()
    # Clean up potential markdown wrapping
    if text.startswith('```'):
        text = text.split('\n', 1)[1] if '\n' in text else text[3:]
        if text.endswith('```'):
            text = text[:-3]
        text = text.strip()

    try:
        data = json.loads(text)
        data['_source_file'] = filename
        return data
    except json.JSONDecodeError:
        print(f"  WARNING: Could not parse response for {filename}")
        return {
            '_source_file': filename,
            'vendor_name': 'PARSE ERROR',
            'notes': text[:200],
        }


def process_all(service, folder_id, subfolder=None):
    """Process all invoices and return extracted data."""
    files = list_files(service, folder_id, subfolder)
    if not files:
        return []

    # Create month folder for filing processed invoices
    month_folder_id, month_folder_name = get_or_create_month_folder(service, folder_id)
    print(f"Filing processed invoices to: {month_folder_name}")

    # Separate duplicates
    seen = set()
    unique_files = []
    duplicate_files = []
    for f in files:
        key = (f['name'], f.get('size', ''))
        if key not in seen:
            seen.add(key)
            unique_files.append(f)
        else:
            duplicate_files.append(f)

    if duplicate_files:
        print(f"  Found {len(duplicate_files)} duplicates - will move to duplicates folder")

    client = anthropic.Anthropic()
    results = []

    for i, f in enumerate(unique_files):
        name = f['name']
        mime = f['mimeType']
        print(f"  [{i+1}/{len(unique_files)}] {name}...", end=" ", flush=True)

        try:
            data = download_file(service, f)

            if mime == 'application/pdf':
                images = pdf_to_images(data)
                if not images:
                    print("SKIP (empty PDF)")
                    continue
                # Process first page (main invoice info is usually there)
                result = extract_invoice_data(client, images[0], 'image/png', name)
            else:
                result = extract_invoice_data(client, data, mime, name)

            results.append(result)
            vendor = result.get('vendor_name', '?')
            total = result.get('total_amount', '?')

            # Move file to month folder
            move_file_to_folder(service, f['id'], month_folder_id)
            print(f"OK - {vendor} - {total}")

        except Exception as e:
            print(f"ERROR: {e}")
            results.append({'_source_file': name, 'vendor_name': 'ERROR', 'notes': str(e)})

    # Move duplicates to a subfolder inside the month folder
    if duplicate_files:
        dup_folder_id, dup_folder_name = get_or_create_month_folder(service, month_folder_id)
        # Override - create "כפילויות" folder, not another month folder
        dup_results = service.files().list(
            q=f"'{month_folder_id}' in parents and mimeType='application/vnd.google-apps.folder' "
              f"and name='כפילויות' and trashed=false",
            fields="files(id)"
        ).execute()
        dup_folders = dup_results.get('files', [])
        if dup_folders:
            dup_folder_id = dup_folders[0]['id']
        else:
            metadata = {
                'name': 'כפילויות',
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [month_folder_id],
            }
            dup_folder = service.files().create(body=metadata, fields='id').execute()
            dup_folder_id = dup_folder['id']

        for f in duplicate_files:
            move_file_to_folder(service, f['id'], dup_folder_id)
            print(f"  Moved duplicate: {f['name']} -> {month_folder_name}/כפילויות")

    return results


def save_excel(results, output_path):
    """Save results to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Installing openpyxl...")
        subprocess.run(['pip3', 'install', 'openpyxl'], capture_output=True)
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.sheet_view.rightToLeft = True  # RTL for Hebrew

    headers = [
        'Vendor', 'Vendor ID', 'Invoice #', 'Date',
        'Description', 'Before VAT', 'VAT', 'Total', 'Currency',
        'Payment', 'Notes'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    fields = [
        'vendor_name', 'vendor_id', 'invoice_number',
        'invoice_date', 'description', 'amount_before_vat', 'vat_amount',
        'total_amount', 'currency', 'payment_method', 'notes'
    ]

    for row_idx, result in enumerate(results, 2):
        for col_idx, field in enumerate(fields, 1):
            value = result.get(field, '')
            if value is None:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'general')

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Summary row
    last_row = len(results) + 2
    ws.cell(row=last_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=5).border = thin_border

    for col in [6, 7, 8]:  # Before VAT, VAT, Total
        formula = f"=SUM({chr(64+col)}2:{chr(64+col)}{last_row-1})"
        cell = ws.cell(row=last_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.border = thin_border

    wb.save(output_path)
    print(f"\nSaved to: {output_path}")


def email_excel(local_path, results):
    """Email the Excel file to the accountant."""
    gmail_user = os.environ['GMAIL_USER']
    gmail_pass = os.environ['GMAIL_APP_PASSWORD']
    to_addr = 'yaffeom@gmail.com'

    now = datetime.now()
    month_name = HEBREW_MONTHS[now.month]
    subject = f'חשבוניות {month_name} {now.year} - {len(results)} חשבוניות'

    total = sum(r.get('total_amount', 0) or 0 for r in results)
    body = f'מצורף קובץ אקסל עם {len(results)} חשבוניות.\nסה"כ: {total:,.2f} ש"ח'

    msg = MIMEMultipart()
    msg['From'] = gmail_user
    msg['To'] = to_addr
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    with open(local_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(local_path)}"')
        msg.attach(part)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(gmail_user, gmail_pass)
        server.sendmail(gmail_user, to_addr, msg.as_string())
    print(f"Emailed Excel to {to_addr}")


def main():
    parser = argparse.ArgumentParser(description="Omer Invoice Reader")
    parser.add_argument("--subfolder", type=str, default=None,
                        help="Process files in a specific subfolder")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel path (default: invoices_YYYYMMDD.xlsx)")
    args = parser.parse_args()

    folder_id = os.environ['DRIVE_FOLDER_ID']

    if not args.output:
        args.output = f"invoices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    print("=" * 50)
    print("OMER INVOICE READER")
    print("=" * 50)

    service = get_drive_service()
    results = process_all(service, folder_id, args.subfolder)

    if results:
        save_excel(results, args.output)
        email_excel(args.output, results)
        print(f"\nProcessed {len(results)} invoices")
    else:
        print("No invoices found to process")


if __name__ == "__main__":
    main()
